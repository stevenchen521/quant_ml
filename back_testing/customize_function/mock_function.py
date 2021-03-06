from collections import OrderedDict
import empyrical as ep
import numpy as np
import matplotlib as mlb
mlb.use("TkAgg")
from pyfolio import timeseries
from pyfolio.utils import (APPROX_BDAYS_PER_MONTH)

import backtrader as bt
import pandas as pd
from backtrader.utils.py3 import items, iteritems
from openpyxl import load_workbook


STAT_FUNCS_PCT = [
    'Annual return',
    'Cumulative returns',
    'Annual volatility',
    'Max drawdown',
    'Daily value at risk',
    'Daily turnover'
]
## rewrite the plot function, add save figure function


################# get the performance matrix
# @patch('pyfolio.plotting.show_perf_stats')
def show_perf_stats_beta(returns, factor_returns=None, positions=None,
                        transactions=None, turnover_denom='AGB',
                        live_start_date=None, bootstrap=False,
                        header_rows=None):
    """
    Prints some performance metrics of the strategy.
    - Shows amount of time the strategy has been run in backtest and
      out-of-sample (in live trading).
    - Shows Omega ratio, max drawdown, Calmar ratio, annual return,
      stability, Sharpe ratio, annual volatility, alpha, and beta.
    Parameters
    ----------
    returns : pd.Series
        Daily returns of the strategy, noncumulative.
         - See full explanation in tears.create_full_tear_sheet.
    factor_returns : pd.Series, optional
        Daily noncumulative returns of the benchmark factor to which betas are
        computed. Usually a benchmark such as market returns.
         - This is in the same style as returns.
    positions : pd.DataFrame, optional
        Daily net position values.
         - See full explanation in create_full_tear_sheet.
    transactions : pd.DataFrame, optional
        Prices and amounts of executed trades. One row per trade.
        - See full explanation in tears.create_full_tear_sheet
    turnover_denom : str, optional
        Either AGB or portfolio_value, default AGB.
        - See full explanation in txn.get_turnover.
    live_start_date : datetime, optional
        The point in time when the strategy began live trading, after
        its backtest period.
    bootstrap : boolean, optional
        Whether to perform bootstrap analysis for the performance
        metrics.
         - For more information, see timeseries.perf_stats_bootstrap
    header_rows : dict or OrderedDict, optional
        Extra rows to display at the top of the displayed table.
    """

    if bootstrap:
        perf_func = timeseries.perf_stats_bootstrap
    else:
        perf_func = timeseries.perf_stats

    perf_stats_all = perf_func(
        returns,
        factor_returns=factor_returns,
        positions=positions,
        transactions=transactions,
        turnover_denom=turnover_denom)

    date_rows = OrderedDict()
    if len(returns.index) > 0:
        date_rows['Start date'] = returns.index[0].strftime('%Y-%m-%d')
        date_rows['End date'] = returns.index[-1].strftime('%Y-%m-%d')

    if live_start_date is not None:
        live_start_date = ep.utils.get_utc_timestamp(live_start_date)
        returns_is = returns[returns.index < live_start_date]
        returns_oos = returns[returns.index >= live_start_date]

        positions_is = None
        positions_oos = None
        transactions_is = None
        transactions_oos = None

        if positions is not None:
            positions_is = positions[positions.index < live_start_date]
            positions_oos = positions[positions.index >= live_start_date]
            if transactions is not None:
                transactions_is = transactions[(transactions.index <
                                                live_start_date)]
                transactions_oos = transactions[(transactions.index >
                                                 live_start_date)]

        perf_stats_is = perf_func(
            returns_is,
            factor_returns=factor_returns,
            positions=positions_is,
            transactions=transactions_is,
            turnover_denom=turnover_denom)

        perf_stats_oos = perf_func(
            returns_oos,
            factor_returns=factor_returns,
            positions=positions_oos,
            transactions=transactions_oos,
            turnover_denom=turnover_denom)
        if len(returns.index) > 0:
            date_rows['In-sample months'] = int(len(returns_is) /
                                                APPROX_BDAYS_PER_MONTH)
            date_rows['Out-of-sample months'] = int(len(returns_oos) /
                                                    APPROX_BDAYS_PER_MONTH)

        perf_stats = pd.concat(OrderedDict([
            ('In-sample', perf_stats_is),
            ('Out-of-sample', perf_stats_oos),
            ('All', perf_stats_all),
        ]), axis=1)
    else:
        if len(returns.index) > 0:
            date_rows['Total months'] = int(len(returns) /
                                            APPROX_BDAYS_PER_MONTH)
        perf_stats = pd.DataFrame(perf_stats_all, columns=['Backtest'])

    for column in perf_stats.columns:
        for stat, value in perf_stats[column].iteritems():
            if stat in STAT_FUNCS_PCT:
                perf_stats.loc[stat, column] = str(np.round(value * 100,
                                                            1)) + '%'
    if header_rows is None:
        header_rows = date_rows
    else:
        header_rows = OrderedDict(header_rows)
        header_rows.update(date_rows)

    return perf_stats
    # utils.print_table(
    #     perf_stats,
    #     float_format='{0:.2f}'.format,
    #     header_rows=header_rows,
    # )


class PerformanceReport():
    """ Report with performce stats for given backtest run
    """
    def __init__(self, stratbt, Indexinfilename, outfilename, commission=None):
        self.stratbt = stratbt  # works for only 1 stategy
        self.Indexinfilename = Indexinfilename
        self.outfilename = outfilename
        self.commission = commission

    def get_startcash(self):
        return self.stratbt.broker.startingcash


    def get_stats(self):
        st = self.stratbt
        pyfoliozer = st.analyzers.getbyname('pyfolio')
        returns, positions, transactions, gross_lev = pyfoliozer.get_pf_items()
        dt = returns.index
        trade_analysis = st.analyzers.myTradeAnalysis.get_analysis()
        rpl = trade_analysis.pnl.net.total
        total_return = rpl / self.get_startcash()
        total_number_trades = trade_analysis.total.total
        trades_closed = trade_analysis.total.closed
        bt_period = dt[-1] - dt[0]
        bt_period_days = bt_period.days
        kpi = {# PnL
               'start_cash': self.get_startcash(),
               'rpl': rpl,
               'result_won_trades': trade_analysis.won.pnl.total,
               'result_lost_trades': trade_analysis.lost.pnl.total,
               'profit_factor': (-1 * trade_analysis.won.pnl.total / trade_analysis.lost.pnl.total),
               'rpl_per_trade': rpl / trades_closed,
               'total_return': 100 * total_return,
               'annual_return': (100 * (1 + total_return)**(365.25 / bt_period_days) - 100),
               # trades
               'total_number_trades': total_number_trades,
               'trades_closed': trades_closed,
               'pct_winning': 100 * trade_analysis.won.total / trades_closed,
               'pct_losing': 100 * trade_analysis.lost.total / trades_closed,
               'avg_money_winning': trade_analysis.won.pnl.average,
               'avg_money_losing':  trade_analysis.lost.pnl.average,
               'best_winning_trade': trade_analysis.won.pnl.max,
               'worst_losing_trade': trade_analysis.lost.pnl.max,
               }
        return kpi

    def generate_excel_report(self, template = None):
        st = self.stratbt
        pyfoliozer = st.analyzers.getbyname('pyfolio')
        returns, positions, transactions, gross_lev = pyfoliozer.get_pf_items()
        ### need to first calculate stats_dict1 becauce return index datetime should be include hour minute and seconds.
        benchmark_rets = get_symbol_returns(filename=self.Indexinfilename)
        stats_dict1 = show_perf_stats_beta(returns, benchmark_rets,
                                          positions=positions,
                                          transactions=transactions)
        stats_dict = self.get_stats()
        stats_dict = pd.DataFrame.from_dict(stats_dict, orient='index')
        stats_dict.rename(columns={0: 'Backtest'}, inplace=True)
        stats_dict = stats_dict.append(stats_dict1)
        stats_dict = stats_dict.round(2)

        # format position
        positions.index.names = ['date']
        positions.index = positions.index.strftime('%Y-%m-%d')
        positions = positions.round(2)

        # format returns
        returns = pd.DataFrame(returns)
        returns.index = returns.index.strftime('%Y-%m-%d')
        returns = returns.round(2)

        ## format transactions df
        transactions.index = transactions.index.strftime('%Y-%m-%d')
        transactions = transactions.round({'price': 2})

        ### get stats_dict run out by pyfolio
        # stats_dict = pd.DataFrame(stats_dict)
        # stats_dict.update(stats_dict1)
        if template== None:
        ### default template
            writer = pd.ExcelWriter(self.outfilename)
            positions.to_excel(writer, sheet_name="backtest_report")
            transactions.to_excel(writer, sheet_name="backtest_report", startcol=5)
            returns.to_excel(writer, sheet_name="backtest_report", startcol=13)
            stats_dict.to_excel(writer, sheet_name="backtest_report", startcol=16)
            writer.save()

        else:
        ### first template
            writer = pd.ExcelWriter(self.outfilename)
            positions.to_excel(writer, sheet_name="trade_positions")
            transactions.to_excel(writer, sheet_name="transactions")
            returns.to_excel(writer, sheet_name="returns")
            stats_dict.to_excel(writer, sheet_name="trade_parameter")
            writer.save()

#
# #### this is just for tp backtest
# class PerformanceReport2(PerformanceReport):
#     def generate_excel_report(self, commission, template=None):
#         st = self.stratbt
#         pyfoliozer = st.analyzers.getbyname('pyfolio')
#         returns, positions, transactions, gross_lev = pyfoliozer.get_pf_items()
#         ### need to first calculate stats_dict1 becauce return index datetime should be include hour minute and seconds.
#         benchmark_rets = get_symbol_returns(filename=self.Indexinfilename)
#         stats_dict1 = show_perf_stats_beta(returns, benchmark_rets,
#                                           positions=positions,
#                                           transactions=transactions)
#         stats_dict = self.get_stats()
#         stats_dict = pd.DataFrame.from_dict(stats_dict, orient='index')
#         stats_dict.rename(columns={0: 'Backtest'}, inplace=True)
#         stats_dict = stats_dict.append(stats_dict1)
#         stats_dict = stats_dict.round(2)
#
#         # format position
#         positions.index.names = ['date']
#         positions.index = positions.index.strftime('%Y-%m-%d')
#         positions = positions.round(2)
#
#         # format returns
#         returns = pd.DataFrame(returns)
#         returns.index = returns.index.strftime('%Y-%m-%d')
#         returns = returns.round(2)
#
#         ## format transactions df
#         transactions.index = transactions.index.strftime('%Y-%m-%d')
#         transactions = transactions.round({'price': 2})
#         transactions['commission'] = abs(transactions['value'] * commission)
#         transactions['action'] = transactions['amount'].apply(lambda x: 'Buy' if x > 0 else 'Sell')
#         # transactions['cost'] = transactions.apply(lambda row: '' if row.value < 0 else)
#
#         # transactions['net'] =
#         # transactions['gross']
#         # transactions['net']
#
#         ### get stats_dict run out by pyfolio
#         # stats_dict = pd.DataFrame(stats_dict)
#         # stats_dict.update(stats_dict1)
#         if template== None:
#         ### default template
#             writer = pd.ExcelWriter(self.outfilename)
#             positions.to_excel(writer, sheet_name="backtest_report")
#             transactions.to_excel(writer, sheet_name="backtest_report", startcol=5)
#             returns.to_excel(writer, sheet_name="backtest_report", startcol=13)
#             stats_dict.to_excel(writer, sheet_name="backtest_report", startcol=16)
#             writer.save()
#         else:
#         ### first template
#             writer = pd.ExcelWriter(self.outfilename)
#             positions.to_excel(writer, sheet_name="trade_positions")
#             transactions.to_excel(writer, sheet_name="transactions")
#             returns.to_excel(writer, sheet_name="returns")
#             stats_dict.to_excel(writer, sheet_name="trade_parameter")
#             writer.save()


# writer.save()
class Cerebro(bt.Cerebro):
    def __init__(self, **kwds):
        super(Cerebro, self).__init__(**kwds)
        self.add_report_analyzers()


    def add_report_analyzers(self):
        """ Adds performance stats, required for report
        """
        self.addanalyzer(bt.analyzers.PyFolio)
        self.addanalyzer(bt.analyzers.DrawDown,
                         _name="myDrawDown")
        self.addanalyzer(bt.analyzers.AnnualReturn,
                         _name="myReturn")
        self.addanalyzer(bt.analyzers.TradeAnalyzer,
                         _name="myTradeAnalysis")


### runstrates is the result that will create after cerebro.run
    def get_strategy_backtest(self):
        return self.runstrats[0][0]



    ## the result input here should be the things run out by cerebro.run()
    def report(self, outfilename, infilename=None, commission=None):
        bt = self.get_strategy_backtest()
        rpt =PerformanceReport(bt, Indexinfilename=infilename, outfilename=outfilename, commission=commission)
        rpt.generate_excel_report()




class BacktestSummary(object):
    def __init__(self, results, input_df, commission, save_path, groupby_list = None):
        self.groupby_list = groupby_list
        self.results = results
        self.input_df = input_df
        self.commission = commission
        self.save_path = save_path


    @staticmethod
    def cal_std_from_returns(df, Buy_date, Sell_date):
        df.Date = pd.to_datetime(df.Date)
        return_df = df[(df.Date >= Buy_date) & (df.Date <= Sell_date)]['close'].pct_change()
        try:
            std = np.nanstd(return_df)
        except Exception:
            std = np.nan
        return std


    @staticmethod
    def return_transactions(txss):
        txs = list()
        # The transactions have a common key (date) and can potentially happend
        # for several assets. The dictionary has a single key and a list of
        # lists. Each sublist contains the fields of a transaction
        # Hence the double loop to undo the list indirection
        for k, v in iteritems(txss):
            for v2 in v:
                txs.append([k] + v2)

        cols = ['date', 'amount', 'price', 'sid', 'symbol', 'value', 'close_type']  # headers are in the first entry
        transactions = pd.DataFrame.from_records(txs, index=range(len(txs)), columns=cols)
        del transactions['sid']
        return transactions

    '''
    if_list = True: return a list. can use apply function to calculate multiple backtesting results
    false: use dataframe output
    '''

    def summary(self, detail_df, if_list):
        value_list = []
        Avg_trans_return = round(detail_df['trans_return'].mean(), 4)
        Avg_daily_return = round(detail_df['daily_return'].mean(), 4)
        Avg_annual_return = round(detail_df['annual_return'].mean(), 4)
        total_period = int(detail_df['period'].sum())
        Avg_period = int(detail_df['period'].mean())
        win_time = len(detail_df[detail_df['trans_return'] > 0])
        loss_time = len(detail_df[detail_df['trans_return'] <= 0])
        total_time = win_time + loss_time
        win_percent = win_time/total_time
        detail_df['percent_value'] = detail_df['trans_return'] + 1
        cum_return = np.prod(detail_df['percent_value']).round(4)
        del detail_df['percent_value']
        try:
            geo_avg_daily_return = round(pow(cum_return, 1 / total_period), 5) - 1
            geo_avg_annual_return = round(pow(geo_avg_daily_return + 1, 365), 4) - 1
        except Exception:
            geo_avg_daily_return = np.nan
            geo_avg_annual_return = np.nan
        daily_risk_free_rate = round(pow(1.03, 1 / 365), 5) - 1
        total_std = np.sum(detail_df['std'] * detail_df['period']) / total_period
        try:
            sharpe_ratio = np.sqrt(365)*(geo_avg_daily_return - daily_risk_free_rate)/total_std
        except Exception:
            sharpe_ratio = np.nan
        if if_list:
            summary = [Avg_trans_return, Avg_daily_return, Avg_annual_return, Avg_period, total_period, cum_return, sharpe_ratio,
                       geo_avg_daily_return, geo_avg_annual_return, win_time, loss_time, total_time, win_percent]
        else:
            summary = pd.DataFrame({'Avg_trans_return': Avg_trans_return,
                                    'Avg_daily_return': Avg_daily_return,
                                    'Avg_annual_return': Avg_annual_return,
                                    'Avg_period': Avg_period,
                                    'total_period': total_period,
                                    'cum_return': cum_return,
                                    'sharpe_ratio': sharpe_ratio,
                                    'geo_avg_daily_return': geo_avg_daily_return,
                                    'geo_avg_annual_return': geo_avg_annual_return,
                                    'win_time': win_time,
                                    'loss_time': loss_time,
                                    'total_time': total_time,
                                    'win_percent': win_percent}, index=[0])
        if self.groupby_list == None:
            pass
        else:
            for ele in self.groupby_list:
                value_ele = detail_df[ele].iloc[0]  ###################################################
                value_list.append(value_ele)
            summary = value_list + summary
        return summary


       # the order of groupby list is important
    '''
    result is the result of Cerebro.run()
    commission rate of transaction
    df is the original backtesting dataframe
    '''
    def format_transaction(self):
        last_close_price = self.input_df['close'].tail(1).iloc[0]
        last_date = pd.to_datetime(self.input_df['Date'].tail(1).iloc[0])
        detail_df = pd.DataFrame()
        end_date = self.input_df.tail(1).Date.values
        # for i in [35]:
        for i in range(len(self.results)):
            if len(self.results) > 1:
                strats = self.results[i][0]
            else:
                strats = self.results[0]
            txss = strats.analyzers.transactions.get_analysis()
            transactions = self.return_transactions(txss)
            Buy_df = transactions[transactions['amount'] > 0]
            if len(Buy_df) == 0:
                continue
            else:
                Buy_df.index = np.arange(len(Buy_df))
                Buy_df.rename(columns={'date': 'Buy_date',
                                       'amount': 'Buy_amount',
                                       'price': 'Buy_price',
                                       'value': 'Buy_value'}, inplace=True)
                del Buy_df['close_type']

            Sell_df = transactions[transactions['amount'] < 0]
            if len(Sell_df) == 0:
                transactions_df = Buy_df
                transactions_df['Sell_date'] = end_date
                transactions_df['Sell_amount'] = -transactions_df['Buy_amount']
                transactions_df['Sell_price'] = last_close_price
                transactions_df['Sell_value'] = last_close_price * -transactions_df['Sell_amount']
                transactions_df['close_type'] = None
            else:
                del Sell_df['symbol']
                Sell_df.index = np.arange(len(Sell_df))
                Sell_df.rename(columns={'date': 'Sell_date',
                                        'amount': 'Sell_amount',
                                        'price': 'Sell_price',
                                        'value': 'Sell_value'}, inplace=True)
                transactions_df = pd.merge(Buy_df, Sell_df, left_index=True, right_index=True, how='left')
                nan_fill = {
                    'Sell_date': last_date,
                    'Sell_amount': -transactions_df['Buy_amount'].iloc[0],
                    'Sell_price': last_close_price,
                    'Sell_value': last_close_price * transactions_df['Buy_amount'].iloc[0],
                    'close_type': 'Not_close'
                }
                transactions_df.fillna(value=nan_fill, inplace=True)

                # del transactions_df['symbol_y']
                # transactions_df.rename(columns={'symbol_x': 'symbol'}, inplace=True)

            # transactions_df['Buy_date'] = transactions_df['Buy_date'].apply(lambda x: x.tz_localize(None))
            # transactions_df['Sell_date'] = transactions_df['Sell_date'].apply(lambda x: x.tz_localize(None))
            transactions_df['std'] = transactions_df.apply(lambda row: self.cal_std_from_returns(self.input_df, row['Buy_date'], row['Sell_date']), axis=1).round(5)
            count = 0
            if self.groupby_list == None:
                pass
            else:
                for ele in self.groupby_list:
                    try:
                        transactions_df[ele] = getattr(strats.params, ele)
                    except Exception:
                        transactions_df[ele] = strats.params.trade_para[count]
                        count = count + 1
            detail_df = detail_df.append(transactions_df, ignore_index=True)

        detail_df['Sell_date'] = detail_df['Sell_date'].values.astype('datetime64[D]')
        detail_df['Buy_date'] = detail_df['Buy_date'].values.astype('datetime64[D]')
        detail_df['period'] = detail_df.apply(lambda row: (row['Sell_date'] - row['Buy_date']).days, axis=1)
        detail_df['Buy_comm'] = (abs(detail_df['Buy_value']) * self.commission).round(2)
        detail_df['Sell_comm'] = (abs(detail_df['Sell_value']) * self.commission).round(2)
        detail_df['net_profit'] = (detail_df['Buy_value'] + detail_df['Sell_value'] - \
                                         detail_df['Buy_comm'] - detail_df['Sell_comm']).round(2)
        detail_df['trans_return'] = (detail_df['net_profit'] / abs(detail_df['Buy_value'])).round(4)
        detail_df['daily_return'] = (pow(detail_df['trans_return'] + 1, 1 / detail_df['period'])).round(5) - 1
        detail_df['annual_return'] = (pow((1 + detail_df['daily_return']), 365)).round(4) - 1
        detail_df['Buy_price'] = detail_df['Buy_price'].round(2)
        detail_df['Sell_price'] = detail_df['Sell_price'].round(2)
        detail_df[['Buy_date', 'Sell_date']] = detail_df[['Buy_date', 'Sell_date']].applymap(lambda n: n.strftime('%Y-%m-%d'))
        # detail_df.rename(columns={'symbol_x': 'symbol'}, inplace=True)
        col_order = ['symbol', 'Buy_amount', 'Buy_price', 'Buy_value', 'Buy_comm', 'Buy_date', 'Sell_amount',
                     'Sell_price', 'Sell_value', 'Sell_comm', 'Sell_date', 'period', 'std', 'net_profit', 'trans_return',
                     'daily_return', 'annual_return', 'close_type']
        if self.groupby_list == None:
            pass
        else:
            col_order = self.groupby_list + col_order
        detail_df = detail_df[col_order]
        return detail_df


    @staticmethod
    def Fetch_raw_data(input_data):
        df1 = pd.read_csv(input_data)
        df1['Date'] = df1['Date'].apply(lambda x: pd.to_datetime(x).strftime("%Y-%m-%d %H:%M:%S"))
        df1.dropna(how="any", inplace=True)
        df1.set_index(['Date'], inplace=True)
        df1['openinterest'] = 0
        col_order = ['open', 'high', 'low', 'close', 'volume', 'openinterest', 'OTri', 'Tri']
        df1 = df1[col_order]
        df1.to_csv(input_data)


    def analyze_and_save_single(self, input_df=True):
        detail_df = self.format_transaction()
        summary_df = self.summary(detail_df, if_list=False)
        df_empty = pd.DataFrame()
        df_empty.to_excel(self.save_path)
        book = load_workbook(self.save_path)
        writer = pd.ExcelWriter(self.save_path, engine='openpyxl')
        writer.book = book
        if input_df:
            self.input_df.to_excel(writer, sheet_name='input_df')
        detail_df.to_excel(writer, sheet_name='detail_transaction')
        summary_df.to_excel(writer, sheet_name='summary')
        writer.save()
        writer.close()


    def analyze_and_save_multi(self, input_df=True):
        detail_df = self.format_transaction()
        summary_df = self.summary(detail_df, if_list=False)
        df_empty = pd.DataFrame()
        df_empty.to_excel(self.save_path)
        book = load_workbook(self.save_path)
        writer = pd.ExcelWriter(self.save_path, engine='openpyxl')
        writer.book = book
        if input_df:
            self.input_df.to_excel(writer, sheet_name='input_df')
        detail_df.to_excel(writer, sheet_name='detail_transaction')
        summary_df.to_excel(writer, sheet_name='summary')
        writer.save()
        writer.close()

