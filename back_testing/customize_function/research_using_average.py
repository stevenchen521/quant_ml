import numpy as np
import pandas as pd
from openpyxl import load_workbook
import pickle
# import seaborn as sns
# from matplotlib import pyplot as plt
# from mpl_toolkits.mplot3d import Axes3D



# def save_as_pickle():
# using mean and delete outlier
def format_summary_excel(in_file, group_list, if_avg_annual):
    def _delete_outlier_format_summary(df, col, group_list):
        four = pd.Series(df[col]).describe()
        Q1 = four['25%']
        Q3 = four['75%']
        IQR = Q3 - Q1
        upper = Q3 + 2 * IQR
        lower = Q1 - 2 * IQR
        value_list = []
        for ele in group_list:
            value_ele = df[ele].iloc[0]  ###################################################
            value_list.append(value_ele)
        df = df[(df[col] <= upper) & (df[col] >= lower)]
        avg = np.nanmean(df[col])
        std = np.nanstd(df[col])
        # median = np.nanmedian(df[col])
        risk_adjusted_return = avg / std
        frequency = len(df)
        avg_period = np.nanmean(df['period'])
        return value_list + [avg, std, risk_adjusted_return, frequency, avg_period]


    x1 = pd.ExcelFile(in_file)
    sheet_name_list = x1.sheet_names
    ticker_list = sheet_name_list[1:]
    count = 0
    summary_list = {}
    for ticker in ticker_list:
        # fig = plt.figure()
        df = pd.read_excel(in_file, sheetname=ticker, usecols=range(20, 41))
        df.index = range(len(df))
        ### rename column name
        for ele in group_list:
            df.rename(columns={ele + '.1': ele}, inplace=True)
        df.dropna(how='any', inplace=True)
        grouped = df.groupby(group_list)
        ### select condition
        if if_avg_annual == True:
            summary_dict = grouped.apply(lambda x: _delete_outlier_format_summary(x, col='annual_return', group_list=group_list), )
        else:
            summary_dict = grouped.apply(lambda x: _delete_outlier_format_summary(x, col='trans_return', group_list=group_list), )
        summary_list[ticker] = summary_dict
        count = count + 1
        print (count)
    return summary_list


def read_pickle(file_name):
    objects = []
    with (open(file_name, "rb")) as openfile:
        while True:
            try:
                objects.append(pickle.load(openfile))
            except EOFError:
                break
    return objects


def format_dict_to_dataframe(groupby_list, pickle_name, columns):
    df = pd.DataFrame()
    objects1 = read_pickle(pickle_name)
    # if if_avg_annual == True:
    #     columns = ['median_avg_annual_return', 'std_avg_annual_return',
    #                 'risk_adjusted_return', 'frequency', 'avg_period']
    #
    # else:
    #     columns = ['median_transaction_return', 'std_transaction_return',
    #                 'risk_adjusted_return', 'frequency', 'avg_period']

    for key in objects1[0].keys():
        summary_dict = objects1[0][key]
        index_list = range(len(summary_dict.apply(lambda x: x[0]).values))
        dict = {}
        m = len(groupby_list)
        for i in range(m):
            dict[groupby_list[i]] = summary_dict.apply(lambda x: x[i]).values
        for l in range(len(columns)):
            dict[columns[l]] = summary_dict.apply(lambda x: x[m + l]).values
        df_columns = groupby_list + columns

        summary_df = pd.DataFrame(data=dict, columns=df_columns, index=index_list)
        # summary_df = pd.DataFrame({'tp_xu':summary_dict.apply(lambda x: x[0]).values,
        #                            'tp_xd': summary_dict.apply(lambda x: x[1]).values,
        #                            'median_transaction_return': summary_dict.apply(lambda x: x[2]).values,
        #                            'std_transaction_return': summary_dict.apply(lambda x: x[3]).values,
        #                            'risk_adjusted_return': summary_dict.apply(lambda x: x[4]).values,
        #                            'frequency': summary_dict.apply(lambda x: x[5]).values,
        #                            'avg_period': summary_dict.apply(lambda x: x[6]).values},
        #                             columns= columns, index=range(len(summary_dict)))
        summary_df['frequency_pct'] = summary_df['frequency']/ summary_df['frequency'].sum()
        summary_df['blp_ticker'] = key
        df = df.append(summary_df)
    return df


def draw_heat_map(df, col, summary_list):
    def _group_func(df, col):
        # blp_ticker = df['blp_ticker'].iloc[0]
        # df.rename(columns={col: blp_ticker}, inplace=True)
        del df['blp_ticker']
        # df.rename(columns={col: blp_ticker}, inplace=True)
        df_copy = df.copy()
        if col == 'frequency' or col == 'avg_period':
            pass
        elif col == 'std_annual_return':
            for columns in df_copy.columns:
                df_copy[columns] = df_copy[columns].rank(method='max', ascending=True)
        else:
            for columns in df_copy.columns:
                df_copy[columns] = df_copy[columns].rank(method='max', ascending=False)
        df_transposed = df_copy.T
        # df.rename(columns={col: blp_ticker}, inplace=True)
        return df_transposed

    df.set_index(summary_list, inplace=True)
    df1 = df[[col, 'blp_ticker', 'stoploss']]
    # df_test = df1[['avg_annual_return', 'blp_ticker']]
    # df_pivot = df1.pivot("blp_ticker", ["tp_xu", "tp_xd"], "avg_annual_return")
    df_for_sns = df1.groupby('blp_ticker', 'stoploss').apply(lambda x: _group_func(x, col))
    # print df_for_sns
    # ax = sns.heatmap(df, annot=True, fmt='d')
    # df_for_sns.to_excel('excel_file\\' + col+'_summary.xlsx')
    return df_for_sns


def format_raw_research_main(group_list, summary_list, in_file='factor_analysis/One_factor_rolling_stoploss.xlsx', pkl_file='rolling_stoploss_research.pkl',
                             path='factor_analysis/rolling_stoploss_research.xlsx', if_pickle_exist=True, if_avg_annual=True):
    if if_pickle_exist == False:
        ## format the research pickle
        df = format_summary_excel(in_file, group_list, if_avg_annual=if_avg_annual)
        output = open(pkl_file, 'wb')
        pickle.dump(df, output)
        output.close()
    else:
        pass
    if if_avg_annual == True:
        col = ['avg_annual_return', 'std_annual_return',
                    'risk_adjusted_return', 'frequency', 'avg_period']
    else:
        col = ['avg_transaction_return', 'std_transaction_return', 'risk_adjusted_return',
               'frequency', 'avg_period']
    # col = ['std_transaction_return', 'risk_adjusted_return', 'frequency', 'avg_period']
    df = format_dict_to_dataframe(group_list, pickle_name=pkl_file, columns=col)
    for col_ele in col:
        # print col_ele
        try:
            book1 = load_workbook(path)
        except Exception:
            df_empty = pd.DataFrame()
            df_empty.to_excel(path)
            book1 = load_workbook(path)
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book = book1
        df_summary = draw_heat_map(df, col_ele, summary_list)
        df_summary.to_excel(writer, sheet_name=col_ele)
        if col_ele == col[0]:
            df.to_excel(writer, sheet_name='raw_data')
        else:
            pass
        writer.save()
        writer.close()


### this is another function
def count_close_type(in_file='data_summary.xlsx', path='close_type_summary.xlsx'):
    x1 = pd.ExcelFile(in_file)
    sheet_name_list = x1.sheet_names
    ticker_list = sheet_name_list[1:]
    count = 0
    summary_df = pd.DataFrame()
    for ticker in ticker_list:
        # fig = plt.figure()
        df = pd.read_excel(in_file, sheetname=ticker, usecols=range(20, 41))
        # df = df[['tp_xu.1', 'tp_windowing.1', 'symbol', 'close_type']]
        summary_df = summary_df.append(df)
    summary_df.to_excel('count_close_type_all.xlsx')


### this is another function
def count_close_type2(in_file='data_summary2.xlsx', path='close_type_summary2.xlsx'):
    x1 = pd.ExcelFile(in_file)
    sheet_name_list = x1.sheet_names
    ticker_list = sheet_name_list[1:]
    count = 0
    summary_df = pd.DataFrame()
    for ticker in ticker_list:
        # fig = plt.figure()
        df = pd.read_excel(in_file, sheetname=ticker)
        # df = df[['tp_xu.1', 'tp_windowing.1', 'symbol', 'close_type']]
        summary_df = summary_df.append(df)
    summary_df.to_excel('count_close_type_all0.6~1.0.xlsx')



if __name__ == '__main__':
    # format_raw_research_main(group_list=['tp_windowing'], summary_list=['tp_windowing'], if_pickle_exist=True,
    #                          path='factor_analysis/rolliavg_ng_stoploss_research2.xlsx',
    #                          pkl_file='rolling_stoploss_research_groupby_tpxu_window_annual_return2.pkl')



    ##### 1.if_pickle_exist: if Ture can jump the format pickle step, otherwise we need to wait 20 mins to format pickle
        # 2.group_list= is the list we want to group and format desired pivot table.
    ##### 3.
    # format_raw_research_main(group_list=['tp_xu', 'tp_windowing', 'stoploss'],
    #                          summary_list=['tp_xu', 'tp_windowing'],
    #                          in_file='data_summary.xlsx',
    #                          if_pickle_exist=True, if_avg_annual=True,
    #                          path='research_result.xlsx',
    #                          pkl_file='summary_result.pkl')

    # format_raw_research_main(group_list=['tp_windowing'], summary_list=['tp_windowing'], if_pickle_exist=False, if_avg_annual= False,
    #                          path='factor_analysis/rolling_stoploss_trans_research2.xlsx',
    #                          pkl_file='rolling_stoploss_research_groupby_tpxu_window_trans_return2.pkl')
    count_close_type2()



