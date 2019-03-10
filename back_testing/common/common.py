import numpy as np
import matplotlib as mlb
mlb.use("TkAgg")
from datetime import datetime
import backtrader as bt
import pandas as pd
from backtrader.utils.py3 import items, iteritems
from openpyxl import load_workbook
import os
import sys


class BacktestSummary(object):
    def __init__(self, domain_name, name , results, input_df, commission, save_path, groupby_list=None):
        self.domain_name = domain_name
        self.groupby_list = groupby_list
        self.results = results
        self.input_df = input_df
        self.commission = commission
        self.save_path = save_path
        self.name = name

    @staticmethod
    def cal_std_from_returns(df, Buy_date, Sell_date):
        df.date = pd.to_datetime(df.date)
        return_df = df[(df.date >= Buy_date) & (df.date <= Sell_date)]['close'].pct_change()
        try:
            std = np.nanstd(return_df)
        except Exception:
            std = np.nan
        return std

    @staticmethod
    def return_transactions(txss):
        txs = list()
        # The transactions have a common key (date) and can potentially happend
        # for several assets. The dictionary has a single key and a list of
        # lists. Each sublist contains the fields of a transaction
        # Hence the double loop to undo the list indirection
        for k, v in iteritems(txss):
            for v2 in v:
                txs.append([k] + v2)

        cols = ['date', 'amount', 'price', 'sid', 'symbol', 'value']  # headers are in the first entry
        transactions = pd.DataFrame.from_records(txs, index=range(len(txs)), columns=cols)
        del transactions['sid']
        return transactions

    '''
    if_list = True: return a list. can use apply function to calculate multiple backtesting results
    false: use dataframe output
    '''

    def summary(self, detail_df, if_list):
        value_list = []
        Avg_trans_return = round(detail_df['trans_return'].mean(), 4)
        Avg_daily_return = round(detail_df['daily_return'].mean(), 4)
        Avg_annual_return = round(detail_df['annual_return'].mean(), 4)
        total_period = int(detail_df['period'].sum())
        Avg_period = int(detail_df['period'].mean())
        win_time = len(detail_df[detail_df['trans_return'] > 0])
        loss_time = len(detail_df[detail_df['trans_return'] <= 0])
        total_time = win_time + loss_time
        win_percent = win_time / total_time
        detail_df['percent_value'] = detail_df['trans_return'] + 1
        cum_return = np.prod(detail_df['percent_value']).round(4)
        del detail_df['percent_value']
        try:
            geo_avg_daily_return = round(pow(cum_return, 1 / total_period), 5) - 1
            geo_avg_annual_return = round(pow(geo_avg_daily_return + 1, 365), 4) - 1
        except Exception:
            geo_avg_daily_return = np.nan
            geo_avg_annual_return = np.nan
        daily_risk_free_rate = round(pow(1.03, 1 / 365), 5) - 1
        total_std = np.sum(detail_df['std'] * detail_df['period']) / total_period
        try:
            sharpe_ratio = np.sqrt(365) * (geo_avg_daily_return - daily_risk_free_rate) / total_std
        except Exception:
            sharpe_ratio = np.nan
        if if_list:
            summary = [Avg_trans_return, Avg_daily_return, Avg_annual_return, Avg_period, total_period, cum_return,
                       sharpe_ratio,
                       geo_avg_daily_return, geo_avg_annual_return, win_time, loss_time, total_time, win_percent]
        else:
            summary = pd.DataFrame({'Avg_trans_return': Avg_trans_return,
                                    'Avg_daily_return': Avg_daily_return,
                                    'Avg_annual_return': Avg_annual_return,
                                    'Avg_period': Avg_period,
                                    'total_period': total_period,
                                    'cum_return': cum_return,
                                    'sharpe_ratio': sharpe_ratio,
                                    'geo_avg_daily_return': geo_avg_daily_return,
                                    'geo_avg_annual_return': geo_avg_annual_return,
                                    'win_time': win_time,
                                    'loss_time': loss_time,
                                    'total_time': total_time,
                                    'win_percent': win_percent}, index=[0])
        if self.groupby_list == None:
            pass
        else:
            for ele in self.groupby_list:
                value_ele = detail_df[ele].iloc[0]  ###################################################
                value_list.append(value_ele)
            summary = value_list + summary
        return summary

    # the order of groupby list is important
    '''
    result is the result of Cerebro.run()
    commission rate of transaction
    df is the original backtesting dataframe
    '''

    def format_transaction(self):
        last_close_price = self.input_df['close'].tail(1).iloc[0]
        last_date = pd.to_datetime(self.input_df['date'].tail(1).iloc[0])
        detail_df = pd.DataFrame()
        end_date = self.input_df.tail(1).date.values
        # for i in [35]:
        for i in range(len(self.results)):
            if len(self.results) > 1:
                strats = self.results[i][0]
            else:
                strats = self.results[0]
            txss = strats.analyzers.transactions.get_analysis()
            transactions = self.return_transactions(txss)
            Buy_df = transactions[transactions['amount'] > 0]
            if len(Buy_df) == 0:
                continue
            else:
                Buy_df.index = np.arange(len(Buy_df))
                Buy_df.rename(columns={'date': 'Buy_date',
                                       'amount': 'Buy_amount',
                                       'price': 'Buy_price',
                                       'value': 'Buy_value'}, inplace=True)
                # del Buy_df['close_type']

            Sell_df = transactions[transactions['amount'] < 0]
            if len(Sell_df) == 0:
                transactions_df = Buy_df
                transactions_df['Sell_date'] = end_date
                transactions_df['Sell_amount'] = -transactions_df['Buy_amount']
                transactions_df['Sell_price'] = last_close_price
                transactions_df['Sell_value'] = last_close_price * -transactions_df['Sell_amount']
                transactions_df['close_type'] = None
            else:
                del Sell_df['symbol']
                Sell_df.index = np.arange(len(Sell_df))
                Sell_df.rename(columns={'date': 'Sell_date',
                                        'amount': 'Sell_amount',
                                        'price': 'Sell_price',
                                        'value': 'Sell_value'}, inplace=True)
                transactions_df = pd.merge(Buy_df, Sell_df, left_index=True, right_index=True, how='left')
                nan_fill = {
                    'Sell_date': last_date,
                    'Sell_amount': -transactions_df['Buy_amount'].iloc[0],
                    'Sell_price': last_close_price,
                    'Sell_value': last_close_price * transactions_df['Buy_amount'].iloc[0]
                    # 'close_type': 'Not_close'
                }
                transactions_df.fillna(value=nan_fill, inplace=True)

                # del transactions_df['symbol_y']
                # transactions_df.rename(columns={'symbol_x': 'symbol'}, inplace=True)

            # transactions_df['Buy_date'] = transactions_df['Buy_date'].apply(lambda x: x.tz_localize(None))
            # transactions_df['Sell_date'] = transactions_df['Sell_date'].apply(lambda x: x.tz_localize(None))
            transactions_df['std'] = transactions_df.apply(
                lambda row: self.cal_std_from_returns(self.input_df, row['Buy_date'], row['Sell_date']), axis=1).round(
                5)
            count = 0
            if self.groupby_list == None:
                pass
            else:
                for ele in self.groupby_list:
                    try:
                        transactions_df[ele] = getattr(strats.params, ele)
                    except Exception:
                        transactions_df[ele] = strats.params.trade_para[count]
                        count = count + 1
            detail_df = detail_df.append(transactions_df, ignore_index=True)

        detail_df['Sell_date'] = detail_df['Sell_date'].values.astype('datetime64[D]')
        detail_df['Buy_date'] = detail_df['Buy_date'].values.astype('datetime64[D]')
        detail_df['period'] = detail_df.apply(lambda row: (row['Sell_date'] - row['Buy_date']).days, axis=1)
        detail_df['Buy_comm'] = (abs(detail_df['Buy_value']) * self.commission).round(2)
        detail_df['Sell_comm'] = (abs(detail_df['Sell_value']) * self.commission).round(2)
        detail_df['net_profit'] = (detail_df['Buy_value'] + detail_df['Sell_value'] - \
                                   detail_df['Buy_comm'] - detail_df['Sell_comm']).round(2)
        detail_df['trans_return'] = (detail_df['net_profit'] / abs(detail_df['Buy_value'])).round(4)
        detail_df['daily_return'] = (pow(detail_df['trans_return'] + 1, 1 / detail_df['period'])).round(5) - 1
        detail_df['annual_return'] = (pow((1 + detail_df['daily_return']), 365)).round(4) - 1
        detail_df['Buy_price'] = detail_df['Buy_price'].round(2)
        detail_df['Sell_price'] = detail_df['Sell_price'].round(2)
        detail_df[['Buy_date', 'Sell_date']] = detail_df[['Buy_date', 'Sell_date']].applymap(
            lambda n: n.strftime('%Y-%m-%d'))
        # detail_df.rename(columns={'symbol_x': 'symbol'}, inplace=True)
        col_order = ['symbol', 'Buy_amount', 'Buy_price', 'Buy_value', 'Buy_comm', 'Buy_date', 'Sell_amount',
                     'Sell_price', 'Sell_value', 'Sell_comm', 'Sell_date', 'period', 'std', 'net_profit',
                     'trans_return', 'daily_return', 'annual_return']
        if self.groupby_list == None:
            pass
        else:
            col_order = self.groupby_list + col_order
        detail_df = detail_df[col_order]
        return detail_df


    @staticmethod
    def Fetch_raw_data(input_data):
        df1 = pd.read_csv(input_data)
        df1['Date'] = df1['Date'].apply(lambda x: pd.to_datetime(x).strftime("%Y-%m-%d %H:%M:%S"))
        df1.dropna(how="any", inplace=True)
        df1.set_index(['Date'], inplace=True)
        df1['openinterest'] = 0
        col_order = ['open', 'high', 'low', 'close', 'volume', 'openinterest', 'OTri', 'Tri']
        df1 = df1[col_order]
        df1.to_csv(input_data)


    def analyze_and_save_single(self, input_df=True):
        detail_df = self.format_transaction()
        summary_df = self.summary(detail_df, if_list=False)
        df_empty = pd.DataFrame()
        df_empty.to_excel(self.save_path)
        book = load_workbook(self.save_path)
        writer = pd.ExcelWriter(self.save_path, engine='openpyxl')
        writer.book = book
        if input_df:
            self.input_df.to_excel(writer, sheet_name='input_df')
        detail_df.to_excel(writer, sheet_name='detail_transaction')
        summary_df.to_excel(writer, sheet_name='summary')
        writer.save()
        writer.close()






class BackTesting(object):
    '''
    Mystrategy: backtrader Mystrategy object;
    input_dict: backtest_data, key is the name of each dataframe
    domain_name: configure strategy dict name

    '''

    def __init__(self, MyStrategy, input_dict, domain_name, target_col, label, save_input_dict=False, commission=0.002, maxcpu=2,
                 initialcash = 10000000):
        self.MyStrategy = MyStrategy
        self.input_dict = input_dict
        self.domain_name = domain_name
        self.commission = commission
        self.maxcpu = maxcpu
        self.target_col = target_col
        self.label = label
        self.initialcash = initialcash
        self.detail_df = pd.DataFrame()
        self.summary_df = pd.DataFrame()
        self.time = datetime.now().strftime("%Y-%m-%d_%H:%M:%S %f")
        self.summary_path = "../../back_testing/summary_excel/{}_summary_{}.xlsx".format(self.domain_name, self.time)
        self.save_input_dict = save_input_dict
        self.default_col = ['date', 'open', 'high', 'low', 'close', 'volume', 'openinterest']
        self.backtest_col = self.default_col + [self.target_col, self.label]

    '''
    input_dict: default, false, if True, save input dict into summary excel
    '''

    def save(self):
        df_empty = pd.DataFrame()
        df_empty.to_excel(self.summary_path)
        book = load_workbook(self.summary_path)
        writer = pd.ExcelWriter(self.summary_path, engine='openpyxl')
        writer.book = book
        self.detail_df.to_excel(writer, sheet_name='detail_transaction')
        self.summary_df.to_excel(writer, sheet_name='summary')
        if self.save_input_dict:
            for key, df in self.input_dict.items():
                sheet_name = key
                df.to_excel(writer, sheet_name=sheet_name)
        writer.save()
        writer.close()



    def backtest(self):
        class PandasData(bt.feeds.PandasData):
            lines = (self.target_col, self.label)
            '''
            The ``dataname`` parameter inherited from ``feed.DataBase`` is the pandas
            DataFrame
            '''
            params = (
                # Possible values for datetime (must always be present)
                #  None : datetime is the "index" in the Pandas Dataframe
                #  -1 : autodetect position or case-wise equal name
                #  >= 0 : numeric index to the colum in the pandas dataframe
                #  string : column name (as index) in the pandas dataframe
                ('datetime', 'date'),

                # Possible values below:
                #  None : column not present
                #  -1 : autodetect position or case-wise equal name
                #  >= 0 : numeric index to the colum in the pandas dataframe
                #  string : column name (as index) in the pandas dataframe
                ('open', -1),
                ('high', -1),
                ('low', -1),
                ('close', -1),
                ('volume', -1),
                ('openinterest', -1),
                (self.target_col, self.target_col),
                (self.label, self.label)
            )


        for key in self.input_dict.keys():
            cerebro = bt.Cerebro(maxcpus =self.maxcpu)
            # Add a strategy
            cerebro.addstrategy(self.MyStrategy)
            # Fetch_raw_data(ticker_data_path)
            dataframe = self.input_dict[key]
            self.unique_col_name = [x for x in list(dataframe.columns) if x not in self.default_col]
            dataframe = dataframe[self.backtest_col]
            dataframe['date'] = dataframe['date'].apply(lambda x: pd.to_datetime(x))
            data = PandasData(dataname=dataframe)
            # Add the Data Feed to Cerebro
            dataname = self.domain_name + "_" + key
            cerebro.adddata(data, name=dataname)
            cerebro.addanalyzer(bt.analyzers.PyFolio)
            cerebro.addanalyzer(bt.analyzers.TradeAnalyzer)
            cerebro.addanalyzer(bt.analyzers.Transactions)
            cerebro.broker.setcash(self.initialcash)
            # cerebro.addsizer(bt.sizers.FixedSize, stake=10000)
            cerebro.broker.setcommission(commission=self.commission)
            print('Starting Portfolio Value: %.2f' % cerebro.broker.getvalue())
            results = cerebro.run()
            '''
            Three input for BecktestSummary:
            results: result of cerebro.run()
            commission rate
            df: the input dataframe, backtesting dataframe
            save_path: the path to save the summary path
            groupby_list: when using optstrategy, add this para, which is the para name list we want to optimize
            '''

            Backtest_summary = BacktestSummary(results=results,
                                               input_df=dataframe,
                                               domain_name=self.domain_name,
                                               name=key,
                                               commission=self.commission,
                                               save_path=self.summary_path,
                                               groupby_list=None)
            detail_df_temp = Backtest_summary.format_transaction()
            summary_df_temp = Backtest_summary.summary(detail_df_temp, if_list=False)
            self.detail_df = self.detail_df.append(detail_df_temp)
            self.summary_df = self.summary_df.append(summary_df_temp)
        self.save()




