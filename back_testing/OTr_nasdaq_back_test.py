# -*- coding: utf-8 -*-
from __future__ import (absolute_import, division, print_function,
                        unicode_literals)
import pandas as pd
import backtrader as bt
from backtrader.feeds import GenericCSVData, DataBase
import datetime
import numpy as np
from openpyxl import load_workbook
from back_testing.customize_function.mock_function import Cerebro
from back_testing.customize_function.customized_analyzer import Transactions, format_transactions
from back_testing.customize_function.Multi_datafeed_test import observers
import os

class GenericCSV_vp(GenericCSVData):
    lines = ('OTri', 'Tri')
    params = (
        ('fromdate', datetime.datetime(2016, 11, 17)),
        ('todate', datetime.datetime(2019, 1, 28)),
        ('OTri', 7),
        ('Tri', 8)
    )


class MyStrategy(bt.Strategy):
    ## trade_para first is tp_xu, second is tp_windowing
    params = (
        ('fromdate', datetime.datetime(2016, 11, 17)),
        ('todate', datetime.datetime(2019, 1, 28))
    )

    def log(self, txt, dt=None):
        ''' Logging function fot this strategy'''
        dt = dt or self.datas[0].datetime.date(0)
        print('%s, %s' % (dt.isoformat(), txt))

    def __init__(self):
        # Keep a reference to the "close" line in the data[0] dataseries
        self.dataclose = self.datas[0].close
        # self.df = pd.DataFrame(columns=['Date', 'Buy_thres', 'Sell_thres', 'Action', 'Price',
        #                                 'Shares', 'Value', 'Commission', 'Gross', 'Net_profit'])
        # To keep track of pending orders and buy price/commission
        self.count = 0
        self.order = None
        self.buyprice = None
        self.buycomm = None
        self.order_dict = {}
        self.startcash = self.broker.getvalue()
        self.close_type = "None"
        # btind.SMA(self.data.tp_score, period=1, subplot=True)
        # btind.SMA(self.data.vp, period=1, subplot=True)
        # btind.SMA(self.data.q_g, period=1, subplot=True)


    def notify(self, order):
        if order.status in [order.Submitted, order.Accepted]:
            return

        if order.status in [order.Completed]:
            if order.isbuy():
                self.log(
                    'BUY EXECUTED, Price: %.2f, Cost: %.2f, Comm %.2f' %
                    (order.executed.price,
                     order.executed.value,
                     order.executed.comm))

                self.buyprice = order.executed.price
                self.buycomm = order.executed.comm
            else:  # Sell
                self.log('SELL EXECUTED, Price: %.2f, Cost: %.2f, Comm %.2f' %
                         (order.executed.price,
                          order.executed.value,
                          order.executed.comm))

            self.bar_executed = len(self)

        elif order.status in [order.Canceled, order.Margin, order.Rejected]:
            self.log('Order Canceled/Margin/Rejected')



    def notify_trade(self, trade):
        if not trade.isclosed:
            return
        self.log('OPERATION PROFIT, GROSS %.2f, NET %.2f' %(trade.pnl, trade.pnlcomm))

        # self.df['Gross'].loc[self.count-1] = trade.pnl
        # self.df['Net_profit'].loc[self.count-1] = trade.pnlcomm

    def next(self):
        self.log('Close, %.2f' % self.dataclose[0])
        if self.order:
            return
        if self.datetime.datetime(ago=0) > datetime.datetime(2016, 11, 18):
            if not self.position: # not in the market
                # Not yet ... we MIGHT BUY if ...
                if (self.data.OTri[0] > 0.5) and (self.data.OTri[-1] < 0.5):
                    # amount_to_invest = (self.p.order_pct * self.broker.cash)
                    # self.size = int(amount_to_invest / self.data.close)
                    self.order = self.buy(size=10)
                    # self.close_type = "None"



## the most import part of rolling stop loss
## compare current close price with the stored one, if current one is greater , replace the stored one

            if self.position:  # in the market
                # Not yet ... we will sell if ...
                if (self.data.OTri[0] < 0.5) and (self.data.OTri[-1] > 0.5):
                    # amount_to_invest = (self.p.order_pct * self.broker.cash)
                    # self.size = int(amount_to_invest / self.data.close)
                    self.order = self.sell(size=10)

                else:
                    pass
            else:
                pass
        else:
            pass


    class BuySellObserver(observers.Observer):
        # alias = ('CashValue',)
        lines = ('buy', 'sell')
        # plotinfo = dict(plot=True, subplot=True)
        def next(self):
            self.lines.buy[0] = self._owner.broker.getcash()
            self.lines.sell[0] = self._owner.broker.getvalue()


    def stop(self):
        pnl = round(self.broker.getvalue() - self.startcash, 2)
        print('Final PnL: {}'.format(pnl))


def cal_std_from_returns(df, Buy_date, Sell_date):
    df.Date = pd.to_datetime(df.Date)
    return_df = df[(df.Date >= Buy_date) & (df.Date <= Sell_date)]['Close'].pct_change()
    try:
        std = np.nanstd(return_df)
    except Exception:
        std = np.nan
    return std


def summary(detail_df, groupby_list):
    value_list = []
    for ele in groupby_list:
        value_ele = detail_df[ele].iloc[0]  ###################################################
        value_list.append(value_ele)
    Avg_trans_return = round(detail_df['trans_return'].mean(), 4)
    Avg_daily_return = round(detail_df['daily_return'].mean(), 4)
    Avg_annual_return = round(detail_df['annual_return'].mean(), 4)
    total_period = int(detail_df['period'].sum())
    Avg_period = int(detail_df['period'].mean())
    win_time = len(detail_df[detail_df['trans_return'] > 0])
    loss_time = len(detail_df[detail_df['trans_return'] <= 0])
    total_time = win_time + loss_time
    win_percent = win_time/total_time
    detail_df['percent_value'] = detail_df['trans_return'] + 1
    cum_return = np.prod(detail_df['percent_value']).round(4)
    del detail_df['percent_value']
    try:
        geo_avg_daily_return = round(pow(cum_return, 1 / total_period), 5) - 1
        geo_avg_annual_return = round(pow(geo_avg_daily_return + 1, 365), 4) - 1
    except Exception:
        geo_avg_daily_return = np.nan
        geo_avg_annual_return = np.nan
    daily_risk_free_rate = round(pow(1.03, 1 / 365), 5) - 1
    total_std = np.sum(detail_df['std'] * detail_df['period']) / total_period
    try:
        sharpe_ratio = np.sqrt(365)*(geo_avg_daily_return - daily_risk_free_rate)/total_std
    except Exception:
        sharpe_ratio = np.nan
    summary = [Avg_trans_return, Avg_daily_return, Avg_annual_return, Avg_period, total_period, cum_return, sharpe_ratio,
               geo_avg_daily_return, geo_avg_annual_return, win_time, loss_time, total_time, win_percent]
    summary = value_list + summary
    return summary


   # the order of groupby list is important
def format_transaction(results, commission, df, groupby_list):
    last_close_price = df['close'].tail(1).iloc[0]
    detail_df = pd.DataFrame()
    end_date = df.tail(1).Date.values
    # for i in [35]:
    for i in range(len(results)):
        print (i)
        # strats = results[i][0]
        strats = results[0]
        txss = strats.analyzers.transactions.get_analysis()
        transactions = format_transactions(txss)
        Buy_df = transactions[transactions['amount'] > 0]
        if len(Buy_df) == 0:
            continue
        else:
            Buy_df.index = np.arange(len(Buy_df))
            Buy_df.rename(columns={'date': 'Buy_date',
                                   'amount': 'Buy_amount',
                                   'price': 'Buy_price',
                                   'value': 'Buy_value'}, inplace=True)
            del Buy_df['close_type']

        Sell_df = transactions[transactions['amount'] < 0]
        if len(Sell_df) == 0:
            transactions_df = Buy_df
            transactions_df['Sell_date'] = end_date
            transactions_df['Sell_amount'] = -transactions_df['Buy_amount']
            transactions_df['Sell_price'] = last_close_price
            transactions_df['Sell_value'] = last_close_price * -transactions_df['Sell_amount']
            transactions_df['close_type'] = None
        else:
            del Sell_df['symbol']
            Sell_df.index = np.arange(len(Sell_df))
            Sell_df.rename(columns={'date': 'Sell_date',
                                    'amount': 'Sell_amount',
                                    'price': 'Sell_price',
                                    'value': 'Sell_value'}, inplace=True)
            transactions_df = pd.merge(Buy_df, Sell_df, left_index=True, right_index=True, how='left')
            nan_fill = {
                'Sell_date': datetime.datetime(2018, 12, 5),
                'Sell_amount': -transactions_df['Buy_amount'].iloc[0],
                'Sell_price': last_close_price,
                'Sell_value': last_close_price * transactions_df['Buy_amount'].iloc[0],
                'close_type': 'Not_close'
            }
            transactions_df.fillna(value=nan_fill, inplace=True)

            # del transactions_df['symbol_y']
            # transactions_df.rename(columns={'symbol_x': 'symbol'}, inplace=True)

        # transactions_df['Buy_date'] = transactions_df['Buy_date'].apply(lambda x: x.tz_localize(None))
        # transactions_df['Sell_date'] = transactions_df['Sell_date'].apply(lambda x: x.tz_localize(None))
        transactions_df['std'] = transactions_df.apply(lambda row: cal_std_from_returns(df, row['Buy_date'], row['Sell_date']), axis=1).round(5)
        count = 0
        if groupby_list == None:
            pass
        else:
            for ele in groupby_list:
                try:
                    transactions_df[ele] = getattr(strats.params, ele)
                except Exception:
                    transactions_df[ele] = strats.params.trade_para[count]
                    count = count + 1
        detail_df = detail_df.append(transactions_df, ignore_index=True)

    detail_df['Sell_date'] = detail_df['Sell_date'].values.astype('datetime64[D]')
    detail_df['Buy_date'] = detail_df['Buy_date'].values.astype('datetime64[D]')
    detail_df['period'] = detail_df.apply(lambda row: (row['Sell_date'] - row['Buy_date']).days, axis=1)
    detail_df['Buy_comm'] = (abs(detail_df['Buy_value']) * commission).round(2)
    detail_df['Sell_comm'] = (abs(detail_df['Sell_value']) * commission).round(2)
    detail_df['net_profit'] = (detail_df['Buy_value'] + detail_df['Sell_value'] - \
                                     detail_df['Buy_comm'] - detail_df['Sell_comm']).round(2)
    detail_df['trans_return'] = (detail_df['net_profit'] / abs(detail_df['Buy_value'])).round(4)
    detail_df['daily_return'] = (pow(detail_df['trans_return'] + 1, 1 / detail_df['period'])).round(5) - 1
    detail_df['annual_return'] = (pow((1 + detail_df['daily_return']), 365)).round(4) - 1
    detail_df['Buy_price'] = detail_df['Buy_price'].round(2)
    detail_df['Sell_price'] = detail_df['Sell_price'].round(2)
    detail_df[['Buy_date', 'Sell_date']] = detail_df[['Buy_date', 'Sell_date']].applymap(lambda n: n.strftime('%Y-%m-%d'))
    # detail_df.rename(columns={'symbol_x': 'symbol'}, inplace=True)
    col_order = ['symbol', 'Buy_amount', 'Buy_price', 'Buy_value', 'Buy_comm', 'Buy_date', 'Sell_amount',
                 'Sell_price', 'Sell_value', 'Sell_comm', 'Sell_date', 'period', 'std', 'net_profit', 'trans_return',
                 'daily_return', 'annual_return', 'close_type']
    col_order = groupby_list + col_order
    detail_df = detail_df[col_order]
    return detail_df


def Fetch_raw_data(input_data):
    df1 = pd.read_csv(input_data)
    df1['Date'] = df1['Date'].apply(lambda x: pd.to_datetime(x).strftime("%Y-%m-%d %H:%M:%S"))
    df1.dropna(how="any", inplace=True)
    df1.set_index(['Date'], inplace=True)
    df1['openinterest'] = 0
    col_order = ['open', 'high', 'low', 'close', 'volume', 'openinterest', 'OTri', 'Tri']
    df1 = df1[col_order]
    df1.to_csv(input_data)


def runstarts():
    import sys
    mypath = os.path.dirname(sys.modules['__main__'].__file__)
    file_name = "nasdaq_for_backtest_processed"
    data_path = mypath + "/data/{}.csv".format(file_name)
    summary_path = mypath + "/summary_excel/{}_summary.xlsx".format(file_name)
    ticker_data_path = data_path
    commission = 0.002
    cerebro = Cerebro(maxcpus=4)
    # Add a strategy
    cerebro.addstrategy(MyStrategy)
    # Fetch_raw_data(ticker_data_path)
    df = pd.read_csv(ticker_data_path)
    data = GenericCSV_vp(dataname=ticker_data_path,)
    # Add the Data Feed to Cerebro
    cerebro.adddata(data, name=file_name)
    cerebro.addanalyzer(bt.analyzers.PyFolio)
    cerebro.addanalyzer(bt.analyzers.TradeAnalyzer)
    cerebro.addanalyzer(Transactions)
    cerebro.broker.setcash(10000000)
    # cerebro.addsizer(bt.sizers.FixedSize, stake=10000)
    # Set the commission
    cerebro.broker.setcommission(commission=commission)
    print('Starting Portfolio Value: %.2f' % cerebro.broker.getvalue())
    # cerebro.addanalyzer(bt.analyzers.SharpeRatio, _name='SharpeRatio')
    # cerebro.addanalyzer(bt.analyzers.DrawDown, _name='DW')
    results = cerebro.run()
    strat = results[0]
    cerebro.plot()
    detail_df = format_transaction(results, 0.002, df, None)
    # df_summary = df_summary.append(detail_df)
    ### save excel part
    try:
        book = load_workbook(summary_path)
    except Exception:
        df_empty = pd.DataFrame()
        df_empty.to_excel(summary_path)
        book = load_workbook(summary_path)
    writer = pd.ExcelWriter(summary_path, engine='openpyxl')
    writer.book = book
    detail_df.to_excel(writer, sheet_name='summary')
    writer.save()
    writer.close()

if __name__ == '__main__':
    runstarts()
